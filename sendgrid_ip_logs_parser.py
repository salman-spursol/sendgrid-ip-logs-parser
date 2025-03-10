"""
SendGrid IP Access Logs parser CLI Application

This script provides a command-line interface for parsing an Excel document
containing IP accesses logs data from SendGrid, and lists those unique IP
addresses from where SendGrid has recently been accessed.
"""

import argparse
import socket
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rich import print as rprint
from rich.console import Console
from rich.table import Table


@dataclass
class ExcelData:
    """Data class to store Excel worksheet information."""

    headers: List[str]
    rows: List[List[Any]]
    sheet_name: str


def resolve_hostname(ip_address: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Convert an IP address to its hostname and FQDN.

    Args:
        ip_address: A string containing an IPv4 or IPv6 address

    Returns:
        A tuple of (hostname, fqdn), where either may be None if resolution fails

    Example:
        >>> resolve_hostname('8.8.8.8')
        ('dns.google', 'dns.google')
    """
    try:
        hostname = socket.gethostbyaddr(ip_address)[0]
        fqdn = socket.getfqdn(ip_address)
        return hostname, fqdn
    except (socket.error, socket.gaierror) as e:
        print(f"ERROR: Failed to resolve {ip_address}: {str(e)}")
        return None, None


def get_host(ip_address: str) -> str:
    return socket.gethostbyaddr(ip_address)[0]


class ExcelParser:
    """Handles the parsing of Excel files."""

    def __init__(self, file_path: Path):
        """Initialize the Excel parser with a file path.

        Args:
            file_path (Path): Path to the Excel file
        """
        self.file_path = file_path
        self.console = Console()
        self.ip_accesses_dict = defaultdict(list)

    def validate_file(self) -> bool:
        """Validate that the file exists and is an Excel file.

        Returns:
            bool: True if file is valid, False otherwise
        """
        if not self.file_path.exists():
            self.console.print(
                f"[red]Error: File {self.file_path} does not exist[/red]"
            )
            return False

        if self.file_path.suffix not in [".xlsx", ".xls", ".xlsm"]:
            self.console.print(
                f"[red]Error: File {self.file_path} is not an Excel file[/red]"
            )
            return False

        return True

    def parse_ip_accesses_logs(self, worksheet: Worksheet) -> int:
        rows_parsed = 0
        for row in worksheet.iter_rows(min_row=2):
            col1_value = row[0].value if row[0].value is not None else ""
            col4_value = row[3].value if row[3].value is not None else ""
            ip_address = col1_value.strip()
            access_method = col4_value.strip()

            rows_parsed += 1
            self.ip_accesses_dict[ip_address].append(access_method)

        return rows_parsed

    def parse_worksheet(self, worksheet: Worksheet) -> ExcelData:
        """Parse a worksheet and extract its data.

        Args:
            worksheet (Worksheet): The worksheet to parse

        Returns:
            ExcelData: Parsed worksheet data
        """
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value if cell.value is not None else ""))

        rows = []
        for row in worksheet.iter_rows(min_row=2):
            print(row)
            row_data = []
            for cell in row:
                row_data.append(cell.value if cell.value is not None else "")
            rows.append(row_data)

        return ExcelData(headers=headers, rows=rows, sheet_name=worksheet.title)

    def display_data(self, data: ExcelData) -> None:
        """Display the parsed Excel data in a formatted table.

        Args:
            data (ExcelData): The parsed Excel data to display
        """
        table = Table(title=f"Sheet: {data.sheet_name}")

        # Add columns
        for header in data.headers:
            table.add_column(header, style="cyan")

        # Add rows
        for row in data.rows:
            table.add_row(*[str(cell) for cell in row])

        self.console.print(table)

    def process_file(self, sheet_name: Optional[str] = None) -> None:
        """Process the Excel file and display its contents.

        Args:
            sheet_name (Optional[str]): Name of specific sheet to process
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path, data_only=True)

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    self.console.print(
                        f"[red]Error: Sheet '{sheet_name}' not found[/red]"
                    )
                    return
                worksheet = workbook[sheet_name]
                # data = self.parse_worksheet(worksheet)
                rows_parsed = self.parse_ip_accesses_logs(worksheet)
                self.console.print(f"{rows_parsed} rows in worksheet processed")
                self.console.print(
                    f"SendGrid has been accessed from {len(self.ip_accesses_dict.keys())} distinct IP addresses:"
                )
                for key in self.ip_accesses_dict.keys():
                    print(key)
                # self.display_data(data)
        except Exception as e:
            self.console.print(f"[red]Error processing file: {str(e)}[/red]")


def main() -> None:
    """Main entry point for the CLI application."""

    # Default fallback filepath
    DEFAULT_EXCEL_PATH = Path(
        r"C:\Users\salman.ahmed\OneDrive - ValueLink Software\Documents\DevSecOps\Production infrastructure\SendGrid_IP_access_logs.xlsx"
    )

    parser = argparse.ArgumentParser(
        description="Parse and display contents of Excel files"
    )
    parser.add_argument(
        "file_path",
        type=Path,
        nargs="?",
        default=DEFAULT_EXCEL_PATH,
        help="Path to the Excel file",
    )

    args = parser.parse_args()

    excel_parser = ExcelParser(args.file_path)

    if not excel_parser.validate_file():
        sys.exit(1)

    excel_parser.process_file("Sheet1")


if __name__ == "__main__":
    main()
